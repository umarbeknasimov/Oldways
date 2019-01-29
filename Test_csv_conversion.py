import os
import glob
import csv
from xlsxwriter.workbook import Workbook as xlsxwriter_wb
for csvfile in glob.glob("*.csv"):
    # saves the order report .xlsx as the same name as the .csv
    # to change the .xlsx file name, replace the text 'csvfile[:-4]' with 'new file name', in quotes
    workbook = xlsxwriter_wb(csvfile[:-4] + '.xlsx') 
    worksheet = workbook.add_worksheet()
    with open(csvfile, 'rt', encoding='utf8') as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader):
            for c, col in enumerate(row):
                try:
                    worksheet.write(r, c, float(col))
                except:
                    worksheet.write(r,c,col)
        workbook.close()
'''

from pyexcel.cookbook import merge_all_to_a_book
# import pyexcel.ext.xlsx # no longer required if you use pyexcel >= 0.2.2 
import glob

for csvfile in glob.glob("*.cvs"):
    merge_all_to_a_book(csvfile, csvfile[:-4]+".xlsx")





import os
import glob
import csv
from openpyxl import Workbook


class FileConvert():
    def __init__(self, filename):
        self.filename = filename

    def convert_csv_to_tsv(self):
        with open(self.filename, "r") as f:
            contents = f.read()
        return contents.replace(",", "\t")

    def save_csv_as_tsv(self, output_path=None):
        filename = self.filename
        contents = self.convert_csv_to_tsv()
        if output_path is None:
            tsv_filename = (filename + ".tsv")
        else:
            tsv_filename = output_path

        print("Saving TSV file as {}".format(tsv_filename))
        with open(tsv_filename, "w") as f:
            f.write(contents)

        return tsv_filename

    def save_tsv_as_xlsx(self, output_path=None, sheet_name="data"):
        filename = self.filename
        with open(filename, "r") as f:
            contents = f.readlines()

        new_filename = (filename + ".xlsx") if output_path is None else output_path

        items = [c.split("\t") for c in contents]
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(sheet_name)
        row_count = 0
        for row in items:
            ws.append(row)
            row_count += 1
            if row_count % 10000 == 0:
                print("processing row {}".format(row_count))


        print("Saving Excel file as {}".format(new_filename))
        wb.save(new_filename)
        return new_filename
s = FileConvert("DefaultOrderExportReport_Jan182019.csv")
s.convert_csv_to_tsv()
s.save_csv_as_tsv()
s.save_tsv_as_xlsx()
'''
'''
import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

f = open('DefaultOrderExportReport_Jan182019.csv', "rU")

csv.register_dialect('colons', delimiter=':')

reader = csv.reader(f, dialect='colons')

wb = Workbook()
dest_filename = r"/tmp/xls_out.xlsx"

ws = wb.worksheets[0]
ws.title = "A Snazzy Title"

for row_index, row in enumerate(reader):
    for column_index, cell in enumerate(row):
        column_letter = get_column_letter((column_index + 1))
        cell = ws['%s%s' % (column_index, row_index)]

wb.save(filename = dest_filename)
'''