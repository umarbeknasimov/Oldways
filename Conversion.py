from openpyxl import Workbook, load_workbook # read and modify Excel 2010 files
from openpyxl.worksheet.dimensions import ColumnDimension # tool to adjust column width

# Convert store order report .csv to .xlsx by reading the .csv cell-by-cell and writing each cell into a .xlsx file
import os
import glob
import csv
from xlsxwriter.workbook import Workbook as xlsxwriter_wb
for csvfile in glob.glob("*.csv"):
    # saves the order report .xlsx as the same name as the .csv
    # to change the .xlsx file name, replace the text 'csvfile[:-4]' with 'new file name', in quotes
    workbook = xlsxwriter_wb(csvfile[:-4] + '.xlsx') 
    worksheet = workbook.add_worksheet()
    with open(csvfile, 'rt', encoding='utf8') as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader):
            for c, col in enumerate(row):
                worksheet.write(r,c,col)
        workbook.close()


# Define the Product class, which represents a single product of which there may be multiple in one order,
# and has certain attributes specific to the product and not to the order
class Product:
    def __init__(self, str_ = None):
        """
        Initializes a product from a comma-spliced string
        str_: string usually in the format "product attribute 1 name: attribute value, product attribute 2 name: attribute value,..." \
            as given in order report column AV
        When we want to artificially make a product that represents shipping costs, str_ will be None
        """
        if str_ is None:
            # If making an artificial product for shipping, skip everything below that sets field values and just return after creating the product
            return
        
        # First, make a dictionary of the fields
        # Result is in the format {field name: field value} for product field names given in default order report column AV
        field_list = [[item.strip() for item in field_value.split(':')]\
                      for field_value in str_.split(',')]
        field_dict = {field: value for field, value in field_list}

        # Set Product object attributes to values stored in the field dictionary
        self.sku = field_dict["Product SKU"]

        # Search the sku-class-item table to find the number of the row that represents this product:
        sku_row = 0
        for i in range(1, sku_sheet.max_row + 1):
            if sku_sheet['C'+str(i)].value == self.sku:
                sku_row = i
                break

        # leave class and item values blank if we don't recognize the sku
        self.class_ = '' if sku_row == 0 else sku_sheet['A'+str(sku_row)].value 
        self.item = '' if sku_row == 0 else sku_sheet['B'+str(sku_row)].value 

        # continue setting product attributes
        self.quantity = int(field_dict["Product Qty"])
        self.unit_cents = round(float(field_dict["Product Unit Price"]) * 100) # used for calculations; in cents
        self.amount = self.unit_cents / 100 # not used for calculations
        self.total_cents = round(float(field_dict["Product Total Price"]) * 100) # in cents

    def from_data(class_ = '', item = '', quantity = 0, unit_cents = 0):
        '''
        Initializes a product based on known attributes (in constrast with __init__, which initializes based on a string read from the order report)
        Should only be used to create an artificial product that represents shipping cost, so usually item = 'POSTAGE & DEL' 
        '''
        product = Product()
        product.class_ = class_
        product.item = item
        product.quantity = quantity
        product.unit_cents = unit_cents
        product.amount = unit_cents / 100
        product.total_cents = quantity * unit_cents
        return product
        

    def write_data(self, ws, row_index):
        """Writes the product-specific data to the worksheet at row_index"""
        items = {"Class": self.class_,
                 "Item": self.item,
                 "Quantity": self.quantity,
                 "Amount": self.amount}
        
        for field, value in items.items():
            ws.cell(row = row_index, column = Order.column_indexes[field]).value = value

# Define the Class (as in SKU/Class/Item) class, which provides functionality to write the "Amount of Sales Receipt" column
class Class:
    def __init__(self, products):
        """
        Initializes a Class object, which contains a list of products in an order that all belong to that class
        products: a list of Product objects that should be in the same class
        """
        self.name = products[0].class_
        self.products = products

    def from_products(products):
        """Converts a list of products to a list of Classes that group the products by class"""
        if len(products) == 0:
            return []
        
        sorted_products = sorted(products, key = lambda x: x.class_) # rearrange list of products so they are sorted by class
        product_list = []
        classes = []
        
        for i, product in enumerate(sorted_products):
            # keep appending current product to current product_list until the class changes, then start a new product_list
            product_list.append(product)
            if i + 1 == len(sorted_products) or sorted_products[i + 1].class_ != product.class_: # Class changed
                classes.append(Class(product_list))
                product_list = []
                
        return classes

    def total_cents(self):
        """Gets the total price of all products in this class in cents"""
        return sum(product.total_cents for product in self.products)

# Define the Order class, which represents one row of the store order report
class Order:
    column_names = ['Customer', 'Date', 'Ref No.', 'Class', 'Payment method', 'Memo', \
                    'Item', 'Quantity', 'Amount', 'Amount of Sales Receipt', 'Amount of transaction', \
                    'Amount Deposited', 'Date deposited to CTC', 'Template Name']
    column_indexes = {name: i + 1 for i, name in enumerate(column_names)}
    
    def __init__(self, ws, row_index):
        """Initializes an order from a worksheet and the row of that order"""

        # First create a dictionary containing column/field names and the corresponding values in row_index
        # Result is in the format {column/field name: value in that column at row row_index}
        # The first row of the order report contains the names of the fields, the row_index row contains the values of the fields
        field_dict = {field: value for field, value in \
                      zip((cell.value for cell in \
                               list(ws.iter_rows(min_row = 1,         max_row = 1)        )[0]), \
                          (cell.value for cell in \
                               list(ws.iter_rows(min_row = row_index, max_row = row_index))[0]))}

        # reformat date from DD-MM-YYYY or DD/MM/YYYY to MM/DD/YYYY
        '''index_of_space = str(field_dict["Order Date"]).find(" ")
        if index_of_space!= -1:
            field_dict["Order Date"] = str(field_dict["Order Date"])[:index_of_space]
        '''
        try: # if date is stored as a datetime object, read the "-" characters and format accordingly
            year_month_date = str(field_dict["Order Date"].date()).split("-")
            field_dict["Order Date"] = year_month_date[2]+"/" \
            +year_month_date[1]+"/"+year_month_date[0]
        except: # otherwise, date is probably stored as a string, so read the "/" characters and format accordingly
            year_month_date = field_dict["Order Date"].split("/")
            field_dict["Order Date"] = year_month_date[1]+"/" \
            +year_month_date[0]+"/"+year_month_date[2]
            
        # From the dictionary of fields and values, set attributes of this Order object
        self.customer = "PRODUCTS"
        self.date = field_dict["Order Date"]
        self.ref_no = field_dict["Customer Name"]
        self.payment = field_dict["Payment Method"]
        self.memo = field_dict["\ufeffOrder ID"]
        self.total_amount = field_dict["Order Total (ex tax)"] # Tax is 0 for all examples given, though
        self.template = "Customer Sales Receipt"
        self.ship_cost = round(float(field_dict["Shipping Cost (ex tax)"]) * 100) # In cents

        # list of products, each initiated from part of the Product Details string, which lists products separated by '|'
        products = [Product(product) for product in field_dict["Product Details"].split('|')]
        self.num_products = len(products)
        # convert list of products to a list of Classes, which each contain a list of products in that class
        self.classes = Class.from_products(products)

        # Add one product to each class that represents shipping cost if there exists a shipping cost
        div = self.ship_cost // len(self.classes)
        mod = self.ship_cost % len(self.classes)
        
        if self.ship_cost != 0:
            # add cost to each class
            for i, class_ in enumerate(self.classes):
                class_.products.append(Product.from_data(class_=class_.name, item="POSTAGE & DEL",\
                                                         quantity=1, unit_cents = div + (1 if i < mod else 0)))

    def total_cents(self):
        '''returns the total price of all products in this order'''
        return sum(class_.total_cents() for class_ in self.classes)

    def write_data(self, ws, row_index):
        """
        Writes the data to the worksheet at row_index, and
        returns the index of the row to insert the next order
        """
        items = {"Customer": self.customer,
                 "Date": self.date,
                 "Ref No.": self.ref_no,
                 "Payment method": self.payment,
                 "Memo": self.memo,
                 "Template Name": self.template}

        curr_row = row_index
        for c, class_ in enumerate(self.classes):
            for product in class_.products:
                # Each row of the same order has the same overall data
                for field, value in items.items():
                    ws.cell(row = curr_row, column = Order.column_indexes[field]).value = value
                # but the products are different
                product.write_data(ws, curr_row)
                
                # a number needs to be added to Ref No. if there are multiple classes
                if len(self.classes) > 1:
                    ws.cell(row = curr_row, column = Order.column_indexes["Ref No."]).value = \
                                self.ref_no + str(c + 1)
                
                curr_row += 1
                
            # only write to "Amount of Sales Receipt" column if the order has more than one class
            if len(self.classes) > 1:
                ws.cell(row = curr_row - 1, column = Order.column_indexes["Amount of Sales Receipt"]).value = \
                            class_.total_cents() / 100
                pass

        # Total amount is on the last line of the order
        ws.cell(row = curr_row - 1, \
                column = Order.column_indexes["Amount of transaction"]).value = self.total_cents() / 100

        return curr_row

# load the order report .xlsx file that we created from the .csv file
loc_order_report = ("DefaultOrderExportReport_Jan182019.xlsx")
wb_order_report = load_workbook(loc_order_report)

# load the .xlsx file containing sku/item/class information
loc_sku_map = ("SKU_class_item.xlsx")
wb_sku_map = load_workbook(loc_sku_map)
sku_sheet = wb_sku_map.worksheets[0]

# create new .xlsx file for formatted sales receipt
wb_new = Workbook()
ws1 = wb_new.active
ws1.title = "Sales Receipts"


# write column names to sales receipt file
for i, column_name in enumerate(Order.column_names):
    ws1.cell(0+1, i+1, column_name)

# read order report rows/orders one-by-one and write corresponding data in sales receipt file
orders = [Order(wb_order_report.active, row) for row in range(2, wb_order_report.active.max_row + 1)]
curr_row = 2
for order in orders:
    curr_row = order.write_data(ws1, curr_row)

#column_dim = ws1.column_dimensions['B']
#column_dim.bestFit = True

wb_new.save("Sales Receipts.xlsx")

'''
wb_new.save("Sales Receipts " \
    + str(datetime.datetime.now().date()) + " " \
    + str(datetime.datetime.now().hour) + ":" + str(datetime.datetime.now().minute) \
    + ".xlsx")
'''
